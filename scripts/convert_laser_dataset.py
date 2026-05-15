"""
Convert Copy of Dataset-Global Laser Based Photo Rejuvenation Devices Market.xlsx
Value sheet -> public/data/value.json and volume.json.

Layout (pivot): row 18 headers; from row 19: indent 0 geography, indent 1 segment type (totals-only in sheet),
indent 2 leaf segment with year columns 2-14 (2021-2033).

Normalization: Excel label 'Other Applications' -> leaf name 'Others'. All rows stay under 'By Application'
(including Hair removal, Skin Tone, Others) so Global rollups match the workbook totals.
"""
from __future__ import annotations

import json
import random
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "Copy of Dataset-Global Laser Based Photo Rejuvenation Devices Market.xlsx"
VALUE_OUT = ROOT / "public/data/value.json"
VOLUME_OUT = ROOT / "public/data/volume.json"

YEARS = list(range(2021, 2034))
HEADER_ROW = 18
DATA_START_ROW = 19

def norm_label(raw: str | None) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if s == "UK":
        return "U.K."
    if s == "Other Applications":
        return "Others"
    if s == "Low/Mid (less then or 6000 US$)":
        return "Low/Mid (~6000 US$)"
    if s == "Premium (More then 6000 US$)":
        return "Premium (More than 6000 US$)"
    return s


def read_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    rows: list[dict] = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        label = cell.value
        if label is None:
            continue
        ind = int(cell.alignment.indent) if cell.alignment and cell.alignment.indent else 0

        year_data: dict[str, float] = {}
        has_any = False
        for i, year in enumerate(YEARS):
            val = ws.cell(row=row_idx, column=2 + i).value
            if val is not None and isinstance(val, (int, float)):
                year_data[str(year)] = float(val)
                has_any = True

        rows.append({
            "row_idx": row_idx,
            "label": norm_label(label),
            "indent": ind,
            "year_data": year_data if has_any else None,
        })
    return rows


def rows_to_nested(rows: list[dict]) -> dict:
    nested: dict = {}
    cur_geo: str | None = None
    cur_seg: str | None = None

    for row in rows:
        lab = row["label"]
        ind = row["indent"]
        yd = row["year_data"]

        if ind == 0:
            cur_geo = lab
            cur_seg = None
            nested.setdefault(cur_geo, {})
            continue

        if cur_geo is None:
            continue

        if ind == 1:
            cur_seg = lab
            continue

        if ind != 2 or yd is None:
            continue

        if not cur_seg:
            continue

        seg_type = cur_seg
        leaf = lab

        geo_bucket = nested[cur_geo].setdefault(seg_type, {})

        geo_bucket[leaf] = {str(yr): round(yd[str(yr)], 1) for yr in YEARS if str(yr) in yd}

    return nested


def generate_volume_from_value(value_data: dict) -> dict:
    """Deterministic pseudo-volume from value (same spirit as legacy convert_excel)."""
    random.seed(42)

    def leaf_factor(base_val: float) -> float:
        if base_val > 10000:
            return random.uniform(400, 800)
        if base_val > 1000:
            return random.uniform(800, 1500)
        return random.uniform(1500, 3000)

    def walk(node):
        if not isinstance(node, dict):
            return node
        year_keys = [k for k in node if str(k).isdigit()]
        dict_children = {k: v for k, v in node.items() if isinstance(v, dict)}

        if year_keys and not dict_children:
            base = next((node[k] for k in year_keys if isinstance(node.get(k), (int, float))), 1)
            factor = leaf_factor(base)
            return {
                k: round(node[k] * factor) if str(k).isdigit() and isinstance(node[k], (int, float)) else node[k]
                for k in node
            }

        if year_keys and dict_children:
            base = next((node[k] for k in year_keys if isinstance(node.get(k), (int, float))), 1)
            factor = leaf_factor(base)
            out = {}
            for k, v in node.items():
                if isinstance(v, dict):
                    out[k] = walk(v)
                elif isinstance(v, (int, float)) and str(k).isdigit():
                    out[k] = round(v * factor)
                else:
                    out[k] = v
            return out

        return {k: walk(v) for k, v in node.items()}

    return walk(value_data)


def main() -> None:
    if not EXCEL_PATH.exists():
        print(f"Missing Excel file: {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Value"]
    rows = read_rows(ws)
    wb.close()

    nested = rows_to_nested(rows)
    geos = list(nested.keys())
    print(f"  Geographies: {len(geos)} — {geos}")

    VALUE_OUT.parent.mkdir(parents=True, exist_ok=True)
    VALUE_OUT.write_text(json.dumps(nested, indent=2), encoding="utf-8")
    print(f"Wrote {VALUE_OUT.relative_to(ROOT)}")

    vol = generate_volume_from_value(nested)
    VOLUME_OUT.write_text(json.dumps(vol, indent=2), encoding="utf-8")
    print(f"Wrote {VOLUME_OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
